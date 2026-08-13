# -*- coding: utf-8 -*-
"""v1.7.3 修复回归：公式 Excel、预检、keep_data_uris、插件列表、stdin。"""

from __future__ import annotations

import io
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from cleanup import clean_markdown  # noqa: E402
from converter import (  # noqa: E402
    ConversionError,
    convert_path,
    convert_stream,
    enforce_precheck,
    is_hard_block_warning,
    precheck_source,
)
from excel_convert import convert_excel_to_markdown  # noqa: E402
from plugin_loader import format_plugin_list, list_plugin_entry_points  # noqa: E402
from zip_convert import sniff_archive_kind  # noqa: E402


class TestExcelFormulaFallback(unittest.TestCase):
    def test_uncached_formula_keeps_formula_text(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "formula.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "EXCEL_FORMULA_MARKER"
            ws["A2"] = 100
            ws["B2"] = 200
            ws["C2"] = "=A2+B2"
            wb.save(path)
            md = convert_excel_to_markdown(path)
            self.assertIn("EXCEL_FORMULA_MARKER", md)
            # 无缓存时至少应保留公式文本，不能整格空白
            self.assertTrue("=A2+B2" in md or "300" in md)


class TestPrecheckHardBlock(unittest.TestCase):
    def test_fake_rar_hard_block(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "fake.zip"
            path.write_bytes(b"Rar!\x1a\x07\x00fake")
            self.assertEqual(sniff_archive_kind(path), "rar")
            tip = precheck_source(str(path))
            self.assertIsNotNone(tip)
            self.assertTrue(is_hard_block_warning(tip))
            with self.assertRaises(ConversionError) as ctx:
                enforce_precheck(str(path))
            self.assertIn("RAR", str(ctx.exception))
            with self.assertRaises(ConversionError) as ctx2:
                convert_path(str(path), local_only=True)
            self.assertIn("RAR", str(ctx2.exception))


class TestKeepDataUrisCleanup(unittest.TestCase):
    def test_clean_markdown_respects_keep_flag(self):
        blob = "A" * 300
        text = f"![pic](data:image/png;base64,{blob})\n"
        truncated = clean_markdown(text, keep_data_uris=False)
        kept = clean_markdown(text, keep_data_uris=True)
        self.assertIn("data:image/...base64...", truncated)
        self.assertIn(blob, kept)
        self.assertNotIn("data:image/...base64...", kept)


class TestConvertStream(unittest.TestCase):
    def test_stream_plain_text(self):
        data = b"hello stdin convert"
        md = convert_stream(io.BytesIO(data), extension=".txt")
        self.assertIn("hello stdin convert", md)


class TestPluginList(unittest.TestCase):
    def test_list_plugins_nonempty_or_ocr_hint(self):
        text = format_plugin_list()
        self.assertIn("插件", text)
        # 开发环境应能探测到 markitdown-ocr；至少函数可调用
        _ = list_plugin_entry_points()


class TestScheduleFixture(unittest.TestCase):
    def test_schedule_fixture_keywords(self):
        sys.path.insert(0, str(ROOT / "tests"))
        from make_schedule_fixture import ensure_schedule_staff_xlsx

        path = ensure_schedule_staff_xlsx(ROOT / "tests" / "samples" / "schedule_staff.xlsx")
        md = convert_path(str(path), local_only=False)
        for kw in ("周一上午", "周一下午", "王成", "杨雨糯", "周五一天", "潘爱萍"):
            self.assertIn(kw, md)


if __name__ == "__main__":
    unittest.main(verbosity=2)
