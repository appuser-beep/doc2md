# -*- coding: utf-8 -*-
"""Excel 合并单元格折叠：排班表 / 宽合并回归测试。"""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from excel_convert import (  # noqa: E402
    _collapse_filled_merges,
    _collapse_wide_merge_row,
    convert_excel_to_markdown,
)

SAMPLES = ROOT / "tests" / "excel_merge" / "samples"
OUTPUT = ROOT / "tests" / "excel_merge" / "output"
REAL_SCHEDULE = Path(r"c:\Users\wluser\Desktop\周一至周五档案整理.xlsx")


def _ensure_schedule_sample() -> Path:
    """生成与用户排班表同构的样例：A 列时段纵合并，B:I 名单横合并。"""
    from openpyxl import Workbook

    SAMPLES.mkdir(parents=True, exist_ok=True)
    path = SAMPLES / "schedule_wide_merge.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "时间"
    ws["B1"] = "周一到周五档案整理人员名单"
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:I2")

    blocks = [
        (3, 4, "周一上午", "无"),
        (5, 7, "周一下午", "王成，刘雨姗，郭家付"),
        (8, 10, "周二上午", "陶迅微"),
        (11, 13, "周二下午", "无"),
        (14, 16, "周三早上", "无"),
        (17, 19, "周三下午", "杨雨糯，张浩，马滟琳"),
        (20, 22, "周四", "无"),
        (23, 27, "周五一天", "刘雨姗，郭家付，屈玉倩（上午），潘爱萍（下午）"),
    ]
    for r1, r2, when, who in blocks:
        ws.cell(r1, 1, when)
        ws.cell(r1, 2, who)
        ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
        ws.merge_cells(start_row=r1, start_column=2, end_row=r2, end_column=9)

    wb.save(path)
    return path


class TestCollapseHelpers(unittest.TestCase):
    def test_wide_merge_row_label_plus_rest(self):
        row = ["周一上午"] + ["无"] * 8
        self.assertEqual(_collapse_wide_merge_row(row), ["周一上午", "无"])

    def test_do_not_collapse_two_equal_metrics(self):
        # 相邻两格碰巧相同，不应误压（阈值 < 3）
        row = ["华北", "100", "100"]
        self.assertEqual(_collapse_wide_merge_row(row), row)

    def test_do_not_collapse_all_equal_survey_row(self):
        # 问卷「是|是|是|是」不得压成一格
        self.assertEqual(_collapse_wide_merge_row(["是", "是", "是", "是"]), ["是", "是", "是", "是"])
        self.assertEqual(_collapse_wide_merge_row(["是", "是", "是"]), ["是", "是", "是"])

    def test_dedupe_vertical_repeats(self):
        grid = [
            ["时间", "名单", "名单", "名单", "名单"],
            ["周一", "甲", "甲", "甲", "甲"],
            ["周一", "甲", "甲", "甲", "甲"],
            ["周二", "乙", "乙", "乙", "乙"],
        ]
        out = _collapse_filled_merges(grid)
        self.assertEqual(out, [["时间", "名单"], ["周一", "甲"], ["周二", "乙"]])

    def test_collapse_pairwise_duplicate_columns(self):
        from excel_convert import _collapse_identical_adjacent_columns

        grid = [
            ["序号", "姓名", "年级专业", "年级专业", "学号", "学号"],
            ["1", "杨燕", "24大数据1班", "24大数据1班", "202442030144", "202442030144"],
            ["2", "李四", "23计科", "23计科", "202342030001", "202342030001"],
        ]
        out = _collapse_identical_adjacent_columns(grid)
        self.assertEqual(
            out,
            [
                ["序号", "姓名", "年级专业", "学号"],
                ["1", "杨燕", "24大数据1班", "202442030144"],
                ["2", "李四", "23计科", "202342030001"],
            ],
        )

    def test_do_not_collapse_distinct_headers_same_values(self):
        from excel_convert import _collapse_identical_adjacent_columns

        grid = [
            ["地区", "指标A", "指标B"],
            ["华北", "100", "100"],
        ]
        self.assertEqual(_collapse_identical_adjacent_columns(grid), grid)


class TestScheduleConversion(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        OUTPUT.mkdir(parents=True, exist_ok=True)
        cls.sample = _ensure_schedule_sample()

    def _assert_clean_schedule(self, md: str, path_label: str):
        out = OUTPUT / f"{path_label}.md"
        out.write_text(md, encoding="utf-8")

        self.assertIn("| 时间 |", md)
        self.assertIn("周一上午", md)
        self.assertIn("周一下午", md)
        self.assertIn("周五一天", md)

        # 不得出现「同一名单横向刷满」：一行里同一名单不应重复出现
        for line in md.splitlines():
            if not line.startswith("|"):
                continue
            if "王成" in line or "陶迅微" in line or "杨雨糯" in line:
                # 单元格内名单本身可含逗号，但不应有「| 王成... | 王成... |」双份
                parts = [p.strip() for p in line.strip("|").split("|")]
                if len(parts) > 2:
                    self.fail(f"名单行仍超过 2 列: {line}")

        # 时段行不应因纵合并重复
        self.assertEqual(md.count("| 周一上午 |"), 1)
        self.assertEqual(md.count("| 周一下午 |"), 1)
        self.assertEqual(md.count("| 周五一天 |"), 1)

        # 列数应为 2（表头 + 分隔 + 数据）
        data_rows = [
            ln for ln in md.splitlines() if ln.startswith("|") and "---" not in ln
        ]
        for ln in data_rows:
            cols = [c.strip() for c in ln.strip().strip("|").split("|")]
            self.assertEqual(len(cols), 2, ln)

    def test_synthetic_schedule(self):
        md = convert_excel_to_markdown(self.sample)
        self._assert_clean_schedule(md, "schedule_wide_merge")

    def test_real_desktop_schedule_if_present(self):
        if not REAL_SCHEDULE.is_file():
            self.skipTest("桌面原件不存在，跳过")
        md = convert_excel_to_markdown(REAL_SCHEDULE)
        self._assert_clean_schedule(md, "real_desktop_schedule")
        self.assertIn("王成，刘雨姗，郭家付，杨进勇，屈玉倩", md)
        self.assertIn("陶迅微", md)
        self.assertNotIn("| 陶迅微 | 陶迅微 |", md)


class TestSideTableRegression(unittest.TestCase):
    def test_ex05_still_splits(self):
        path = ROOT / "tests" / "office_deep" / "samples" / "EX05_multi_table_side.xlsx"
        if not path.is_file():
            self.skipTest("EX05 样例缺失")
        md = convert_excel_to_markdown(path)
        self.assertIn("区域 1", md)
        self.assertIn("区域 2", md)
        self.assertGreaterEqual(md.count("|"), 6)


if __name__ == "__main__":
    unittest.main(verbosity=2)
