"""生成排班表回归样例（门禁 / 单测共用）。"""

from __future__ import annotations

from pathlib import Path


def ensure_schedule_staff_xlsx(path: Path | None = None) -> Path:
    """生成与真实排班表同构的 xlsx，保证门禁关键词可命中。"""
    from openpyxl import Workbook

    if path is None:
        path = Path(__file__).resolve().parent / "samples" / "schedule_staff.xlsx"
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "时间"
    ws["B1"] = "周一到周五档案整理人员名单"
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:I2")

    blocks = [
        (3, 4, "周一上午", "无"),
        (5, 7, "周一下午", "王成，刘雨姗，郭家付"),
        (8, 10, "周二上午", "陶迅微"),
        (11, 13, "周二下午", "无"),
        (14, 16, "周三早上", "无"),
        (17, 19, "周三下午", "杨雨糯，张浩，马滟琳"),
        (20, 22, "周四", "无"),
        (23, 27, "周五一天", "刘雨姗，郭家付，屈玉倩（上午），潘爱萍（下午）"),
    ]
    for r1, r2, when, who in blocks:
        ws.cell(r1, 1, when)
        ws.cell(r1, 2, who)
        ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
        ws.merge_cells(start_row=r1, start_column=2, end_row=r2, end_column=9)

    wb.save(path)
    return path


if __name__ == "__main__":
    p = ensure_schedule_staff_xlsx()
    print(p)
