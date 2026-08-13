"""Excel 增强转换：填充合并单元格、按空白列拆分左右表、输出干净 Markdown。"""

from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path


def _escape_md_cell(text: str) -> str:
    """幂等转义单元格内 |。"""
    s = text or ""
    out: list[str] = []
    i = 0
    while i < len(s):
        if s[i] == "\\" and i + 1 < len(s) and s[i + 1] == "|":
            out.append("\\|")
            i += 2
            continue
        if s[i] == "|":
            out.append("\\|")
            i += 1
            continue
        out.append(s[i])
        i += 1
    return "".join(out)


def _format_number(value: float, number_format: str | None) -> str:
    fmt = (number_format or "General").strip()
    low = fmt.lower()
    if value != value:  # NaN
        return ""
    # 百分比
    if "%" in fmt:
        pct = value * 100
        if abs(pct - round(pct)) < 1e-9:
            return f"{int(round(pct))}%"
        return f"{pct:g}%"
    # 科学计数（格式含 E+ / E- 时强制科学记法，避免 .6g 对中等整数省略指数）
    if "e+" in low or "e-" in low or re.search(r"0\.0*e", low):
        return f"{value:.6E}"

    decimals_m = re.search(r"[0#]\.(0+|#+)", fmt)
    decimals = len(decimals_m.group(1)) if decimals_m else None

    # 货币：¥ ￥ $ € £ 以及 [$¥-804] / [$$-409] 形式
    currency = None
    m = re.search(r"\[\$([^\-\]]+)", fmt)
    if m:
        raw = m.group(1).strip()
        if raw in {"¥", "￥", "\u00a5"}:
            currency = "¥"
        elif raw:
            currency = raw[0] if len(raw) <= 2 else raw
    if currency is None:
        for sym in ("¥", "￥", "\u00a5", "$", "€", "£"):
            if sym in fmt or f'"{sym}"' in fmt:
                currency = "¥" if sym in {"¥", "￥", "\u00a5"} else sym
                break
    if currency:
        d = 2 if decimals is None else decimals
        return f"{currency}{value:,.{d}f}"
    # 千分位
    if "#,##" in fmt or "#,#" in fmt:
        if decimals is None:
            if abs(value - round(value)) < 1e-9:
                return f"{value:,.0f}"
            return f"{value:,.2f}"
        return f"{value:,.{decimals}f}"
    # 固定小数
    if decimals is not None:
        return f"{value:.{decimals}f}"
    if abs(value - int(value)) < 1e-9 and "0.00" not in fmt and ".0" not in low:
        return str(int(value))
    return str(value)


def _cell_str(value, number_format: str | None = None) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        fmt = (number_format or "").lower()
        # 纯日期（无时间或时间全 0）且格式不像时间
        if (
            value.hour == 0
            and value.minute == 0
            and value.second == 0
            and value.microsecond == 0
            and not any(tok in fmt for tok in ("h", "时", "分", "s", "am", "pm"))
        ):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float):
        return _escape_md_cell(_format_number(value, number_format))
    if isinstance(value, int) and not isinstance(value, bool):
        return _escape_md_cell(_format_number(float(value), number_format))
    s = str(value).strip()
    if s.lower() == "nan":
        return ""
    return _escape_md_cell(s.replace("\n", "<br>"))


def _pick_cell_value(cached, formula, number_format: str | None = None) -> str:
    """优先用已缓存的计算值；没有缓存时保留公式文本，避免整格变空。"""
    cached_s = _cell_str(cached, number_format)
    if cached_s:
        return cached_s
    return _cell_str(formula, number_format)


def _load_grid(path: Path) -> list[tuple[str, list[list[str]]]]:
    """返回 [(sheet_name, grid), ...]，grid 已填充合并单元格。"""
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries

    # data_only=True 在无 Excel 缓存时公式格为 None；再读一份公式本底作回退
    wb_cached = load_workbook(path, data_only=True)
    wb_formula = load_workbook(path, data_only=False)
    sheets = []
    for ws_c, ws_f in zip(wb_cached.worksheets, wb_formula.worksheets):
        max_r = max(ws_c.max_row or 0, ws_f.max_row or 0)
        max_c = max(ws_c.max_column or 0, ws_f.max_column or 0)
        if max_r == 0 or max_c == 0:
            continue
        grid = [["" for _ in range(max_c)] for _ in range(max_r)]
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell_f = ws_f.cell(r, c)
                grid[r - 1][c - 1] = _pick_cell_value(
                    ws_c.cell(r, c).value,
                    cell_f.value,
                    cell_f.number_format,
                )

        for merged in ws_f.merged_cells.ranges:
            min_c, min_r, max_c2, max_r2 = range_boundaries(str(merged))
            val = grid[min_r - 1][min_c - 1]
            for r in range(min_r, max_r2 + 1):
                for c in range(min_c, max_c2 + 1):
                    grid[r - 1][c - 1] = val
        sheets.append((ws_f.title, grid))
    return sheets


def _column_nonempty_count(grid: list[list[str]], col: int) -> int:
    return sum(1 for row in grid if col < len(row) and row[col].strip())


def _split_regions(grid: list[list[str]]) -> list[list[list[str]]]:
    """按「整列几乎为空」把并排区域拆开。"""
    if not grid:
        return []
    width = max(len(r) for r in grid)
    # 补齐
    grid = [r + [""] * (width - len(r)) for r in grid]

    nonempty = [_column_nonempty_count(grid, c) for c in range(width)]
    max_ne = max(nonempty) if nonempty else 0
    # 分隔列必须「几乎全空」。
    # 旧逻辑用 max_ne//10，会把「仅表头+一行有值、下方大量空白」的宽表列误判为分隔，导致丢列。
    blank_cols: set[int] = set()
    for c, n in enumerate(nonempty):
        if n == 0:
            blank_cols.add(c)
        elif max_ne >= 10 and n <= 1 and n <= max_ne // 20:
            # 峰值很高时，仅 1 个噪点的空隙列仍视为分隔（并排多表场景）
            blank_cols.add(c)

    regions: list[list[list[str]]] = []
    start = None
    for c in range(width):
        if c in blank_cols:
            if start is not None:
                regions.append([row[start:c] for row in grid])
                start = None
        else:
            if start is None:
                start = c
    if start is not None:
        regions.append([row[start:width] for row in grid])

    # 去掉全空区域
    cleaned = []
    for reg in regions:
        if any(any(cell.strip() for cell in row) for row in reg):
            cleaned.append(reg)
    return cleaned or [grid]


def _trim_grid(grid: list[list[str]]) -> list[list[str]]:
    if not grid:
        return grid
    # 去尾部空行
    while grid and not any(c.strip() for c in grid[-1]):
        grid = grid[:-1]
    # 去尾部空列
    if not grid:
        return grid
    width = len(grid[0])
    while width > 0 and all(len(r) <= width - 1 or not r[width - 1].strip() for r in grid):
        width -= 1
        grid = [r[:width] for r in grid]
    # 去全空列（中间）
    keep = []
    for c in range(width):
        if any(r[c].strip() for r in grid if c < len(r)):
            keep.append(c)
    grid = [[r[c] for c in keep] for r in grid]
    # 去全空行
    grid = [r for r in grid if any(c.strip() for c in r)]
    return grid


def _collapse_wide_merge_row(row: list[str]) -> list[str]:
    """把「合并单元格填充」造成的整行重复压回逻辑列。

    例：[时间段, 名单, 名单, 名单, ...] → [时间段, 名单]
    避免值班表类 Excel 转成 8 列重复垃圾表。
    """
    if len(row) <= 2:
        return list(row)
    first, rest = row[0], row[1:]
    # 首列标签 + 右侧全部相同（典型：A 列时间，B:I 合并人员）
    if len(rest) >= 3 and len(set(rest)) == 1:
        return [first, rest[0]]
    # 整行同一内容（宽标题横幅）
    nonempty = [c for c in row if c.strip()]
    if len(nonempty) >= 3 and len(set(nonempty)) == 1:
        return [nonempty[0]]
    return list(row)


def _collapse_identical_adjacent_columns(grid: list[list[str]]) -> list[list[str]]:
    """折叠整表中「相邻列内容完全相同」的合并填充列。

    例：| 年级专业 | 年级专业 | 学号 | 学号 | → | 年级专业 | 学号 |
    表头不同则不折叠（避免把碰巧同值的两指标列压掉）。
    """
    if not grid:
        return grid
    width = max((len(r) for r in grid), default=0)
    if width <= 1:
        return [list(r) for r in grid]
    padded = [list(r) + [""] * (width - len(r)) for r in grid]
    keep: list[int] = []
    i = 0
    while i < width:
        j = i + 1
        while j < width and all(
            padded[r][i].strip() == padded[r][j].strip() for r in range(len(padded))
        ):
            j += 1
        keep.append(i)
        i = j
    return [[padded[r][c] for c in keep] for r in range(len(padded))]


def _collapse_filled_merges(grid: list[list[str]]) -> list[list[str]]:
    """折叠横向重复列，并去掉纵向合并导致的连续重复行。"""
    if not grid:
        return grid
    collapsed = [_collapse_wide_merge_row(r) for r in grid]
    width = max((len(r) for r in collapsed), default=0)
    collapsed = [r + [""] * (width - len(r)) for r in collapsed]
    collapsed = _collapse_identical_adjacent_columns(collapsed)
    out: list[list[str]] = []
    for r in collapsed:
        if out and r == out[-1]:
            continue
        out.append(r)
    return out


def _guess_title(grid: list[list[str]]) -> tuple[str | None, list[list[str]]]:
    """若首行像标题（多列相同或仅一格有字），抽成标题。"""
    if not grid:
        return None, grid
    first = grid[0]
    nonempty = [c for c in first if c.strip()]
    if not nonempty:
        return None, grid[1:]
    # 合并标题：多列值相同
    if len(set(nonempty)) == 1 and (len(nonempty) >= 2 or len(first) >= 3):
        return nonempty[0], grid[1:]
    # 单格标题且很短行
    if len(nonempty) == 1 and len(first) > 1 and not any(
        first[i].strip() for i in range(1, len(first))
    ):
        return nonempty[0], grid[1:]
    return None, grid


def _peel_banner_rows(grid: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    """连续剥掉「整行合并/重复」的横幅说明，避免当成表头。"""
    notes: list[str] = []
    body = grid
    while body:
        title, rest = _guess_title(body)
        if title is None:
            # 也可能是「项目名称：」这种仅首列有字、其余空
            first = body[0]
            nonempty = [c for c in first if c.strip()]
            if len(nonempty) == 1 and len(first) >= 2:
                # 若下一行才是真正表头（含序号等），则本行当说明
                header_keys = ("序号", "评分", "权重", "指标", "标准", "名称")
                if len(body) > 1 and sum(
                    1 for c in body[1] if any(k in c for k in header_keys)
                ) >= 2:
                    notes.append(nonempty[0])
                    body = body[1:]
                    continue
            break
        notes.append(title)
        body = rest
    return notes, body


def _grid_to_md_table(grid: list[list[str]]) -> str:
    grid = _trim_grid(grid)
    if not grid:
        return ""
    # 先折叠合并填充造成的重复，再剥横幅
    grid = _collapse_filled_merges(grid)
    grid = _trim_grid(grid)
    notes, body = _peel_banner_rows(grid)
    body = _collapse_filled_merges(body)
    body = _trim_grid(body)
    parts: list[str] = []
    # 第一条横幅常作标题，其余作说明
    if notes:
        parts.append(f"### {notes[0]}")
        for n in notes[1:]:
            parts.append(n)
    if not body:
        return "\n\n".join(parts)

    # 两列「时间 | 内容」类排班表：直接输出，避免假多列表头
    width = max(len(r) for r in body)
    if width <= 1:
        for row in body:
            cell = next((c for c in row if c.strip()), "")
            if cell:
                parts.append(cell)
        return "\n\n".join(parts)

    if width == 2:
        # 若首行像「时间 / 标题」且后续都是时段+人员，用首行作表头
        header = body[0]
        data = body[1:]
        # 表头两格不同且后面行首列像时段标签时，保留表头
        lines = [
            "| " + " | ".join(h if h.strip() else f"列{i+1}" for i, h in enumerate(header)) + " |",
            "| --- | --- |",
        ]
        for row in data:
            row = row + [""] * (2 - len(row))
            if not any(c.strip() for c in row):
                continue
            lines.append("| " + " | ".join(row[:2]) + " |")
        parts.append("\n".join(lines))
        return "\n\n".join(parts)

    header = body[0]
    data = body[1:]
    header_keys = ("序号", "评分", "权重", "指标", "标准", "名称", "项目")
    if data:
        next_hits = sum(1 for c in data[0] if any(k in c for k in header_keys))
        cur_hits = sum(1 for c in header if any(k in c for k in header_keys))
        if next_hits > cur_hits and next_hits >= 2:
            note = " ".join(c for c in header if c.strip())
            if note:
                parts.append(note)
            header = data[0]
            data = data[1:]

    width = max(len(header), max((len(r) for r in data), default=0))
    header = header + [""] * (width - len(header))
    header = [h if h.strip() else f"列{i+1}" for i, h in enumerate(header)]

    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for row in data:
        row = row + [""] * (width - len(row))
        if [c.strip() for c in row] == [c.strip() for c in header]:
            continue
        lines.append("| " + " | ".join(row) + " |")
    parts.append("\n".join(lines))
    return "\n\n".join(parts)


def _append_excel_images(path: Path, markdown: str) -> str:
    """从 xlsx 包内列出嵌入图片，在 Markdown 中标注（便于大模型知晓有图）。"""
    import zipfile

    try:
        with zipfile.ZipFile(path) as zf:
            media = [n for n in zf.namelist() if n.startswith("xl/media/")]
    except Exception:
        return markdown
    if not media:
        return markdown
    lines = [markdown.rstrip(), "", "### 嵌入图片", ""]
    for i, name in enumerate(media, 1):
        lines.append(f"![Excel图片 {i} · {Path(name).name}](embedded-image)")
    return "\n".join(lines) + "\n"


def convert_excel_to_markdown(path: str | Path) -> str:
    path = Path(path)
    if path.suffix.lower() == ".xls":
        # 旧版 xls：openpyxl 不支持，用 pandas 读后走清理逻辑
        import pandas as pd

        frames = pd.read_excel(path, sheet_name=None, header=None, dtype=object)
        out_parts = []
        for name, df in frames.items():
            grid = [[_cell_str(v) for v in row] for row in df.values.tolist()]
            regions = _split_regions(grid)
            bits = []
            for idx, reg in enumerate(regions, 1):
                md = _grid_to_md_table(reg)
                if not md.strip():
                    continue
                bits.append(f"## 区域 {idx}\n\n{md}" if len(regions) > 1 else md)
            if not bits:
                continue
            if len(frames) > 1 or len(regions) > 1:
                out_parts.append(f"# {name}\n\n" + "\n\n".join(bits))
            else:
                out_parts.append(bits[0])
        return "\n\n".join(out_parts).strip() + "\n"

    sheets = _load_grid(path)
    if not sheets:
        return _append_excel_images(path, "")

    out_parts: list[str] = []
    for sheet_name, grid in sheets:
        regions = _split_regions(grid)
        sheet_bits = []
        for idx, reg in enumerate(regions, 1):
            md = _grid_to_md_table(reg)
            if not md.strip():
                continue
            if len(regions) > 1:
                sheet_bits.append(f"## 区域 {idx}\n\n{md}")
            else:
                sheet_bits.append(md)
        if not sheet_bits:
            continue
        if len(sheets) > 1:
            out_parts.append(f"# {sheet_name}\n\n" + "\n\n".join(sheet_bits))
        else:
            if len(regions) > 1:
                out_parts.append(f"# {sheet_name}\n\n" + "\n\n".join(sheet_bits))
            else:
                out_parts.append(sheet_bits[0])
    result = "\n\n".join(out_parts).strip() + "\n"
    return _append_excel_images(path, result)
